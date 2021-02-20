import sys
from openpyxl import Workbook
from openpyxl import load_workbook

def sort_columns(infile, outfile, columns):
    # Load Excel file
    wb = load_workbook(filename=infile)

    # Select active worksheet
    ws = wb.active

    # Insert column and move data
    j = 1
    for i in columns:
        ws.insert_cols(j)
        for k in range (1, 1000):
            data = ws.cell(row = k, column = j + int(i)).value
            ws.cell(row = k, column = j).value = data
        j = j + 1
    columns.sort(reverse=True)
    for i in columns:
        ws.delete_cols(int(i) + j - 1)

    # Save the file
    wb.save(outfile)

# value[0]  = Program name
# value[1]  = Input filename
# value[2]  = Output filename
# value[3:] = Sort index list
if __name__ == '__main__':
    value = sys.argv
    if len(value) > 3:
        infile = value[1]
        outfile = value[2]
        columns = value[3:]
        sort_columns(infile, outfile, columns)
    else:
        print ("argument error")
